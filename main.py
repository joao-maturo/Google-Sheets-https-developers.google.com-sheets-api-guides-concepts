from openpyxl import load_workbook

try:
    # Carrega o arquivo Excel
    workbook = load_workbook(".vscode\\desafio_tuskan\\produtos.xlsx")
    sheet = workbook.active

    # Encontra a próxima linha disponível na coluna do nome
    next_row = sheet.max_row + 1

    # Adiciona os cabeçalhos das provas e situação
    sheet.cell(row=1, column=2, value="prova1")
    sheet.cell(row=1, column=3, value="prova2")
    sheet.cell(row=1, column=4, value="prova3")
    sheet.cell(row=1, column=5, value="média")
    sheet.cell(row=1, column=6, value="Situação")

    # Solicita o número de pessoas
    num_pessoas = int(input("Digite o número de pessoas que deseja adicionar: "))

    for _ in range(num_pessoas):
        # Solicitar o nome do aluno
        nome = input("Digite o nome do aluno: ")

        # Adicionar o nome na próxima linha disponível na coluna do nome
        sheet.cell(row=next_row, column=1, value=nome)

        # Solicitar as notas das provas
        notas = []
        for i in range(3):
            nota = float(input(f"Digite a nota da prova {i+1}: "))
            notas.append(nota)

        # Calcular a média das notas
        media = sum(notas) / 3

        # Determinar a situação
        if media < 5:
            situacao = "Reprovado por Nota"
        elif 5 <= media < 7:
            situacao = "Exame Final"
        else:
            situacao = "Aprovado"

        # Adiciona as notas das provas e a média na mesma linha que o nome
        for col, nota in enumerate(notas, start=2):
            sheet.cell(row=next_row, column=col, value=nota)
        sheet.cell(row=next_row, column=5, value=media)
        sheet.cell(row=next_row, column=6, value=situacao)

        # Avança para a próxima linha disponível
        next_row += 1

    # Salva as alterações no arquivo Excel
    workbook.save(".vscode\\desafio_tuskan\\produtos.xlsx")
    print("Dados salvos com sucesso.")

except Exception as e:
    print("Ocorreu um erro ao salvar os dados:", e)
